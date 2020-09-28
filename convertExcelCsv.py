import openpyxl, csv, os

os.chdir(r'C:\Users\SOLO\Documents\PythonFiles')


os.makedirs('convertedExcels', exist_ok = True)

for excelFilename in os.listdir('.'):
    if not excelFilename.endswith('.xlsx'):
        continue
    print('Loading %s...' %(excelFilename))
    wb = openpyxl.load_workbook(excelFilename)
    for sheetName in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheetName)
        csvFilename = excelFilename[0:len(excelFilename)-5]+ \
                      '_'+sheetName+'.csv'
        print('Converting %s to %s...' % (excelFilename, csvFilename))
        csvFile = open(os.path.join('convertedExcels', csvFilename), \
                       'w', newline='')
        csvWriter = csv.writer(csvFile)
        for rowNum in range(1, sheet.max_row+1):
            rowData = []
            for colNum in range(1, sheet.max_column+1):
                rowData.append(str(sheet.cell(rowNum, colNum).value))
            csvWriter.writerow(rowData)
        csvFile.close()
        
