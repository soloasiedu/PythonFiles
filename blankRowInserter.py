import openpyxl,os,sys


os.chdir(r'C:\Users\SOLO\Documents\PythonFiles')

n = int(sys.argv[1])
m = int(sys.argv[2])
filename = str(sys.argv[-1])
print(n)
print(m)
print(filename)

wb = openpyxl.load_workbook(filename)
sheet = wb.active
nwb = openpyxl.Workbook()
nsheet = nwb.active
print('Copying first rows...')
for r in range(1, n):
    for c in range(1, sheet.max_column+1):
        nsheet.cell(row=r, column=c).value = sheet.cell(row=r, column=c).value

print('Copying remaining rows...')
for r in range(n+m, sheet.max_row):
    for c in range(1, sheet.max_column+1):
        nsheet.cell(row=r, column=c).value = sheet.cell(row=r, column=c).value

print('Done.')
nwb.save('produceSalesRowsInsert.xlsx')


