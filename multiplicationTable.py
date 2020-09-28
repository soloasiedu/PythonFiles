import openpyxl,sys,os
os.chdir(r'C:\Users\SOLO\Documents\PythonFiles')
print(os.getcwd())
n = int(sys.argv[-1])
wb = openpyxl.Workbook()
sheet = wb.active
print(n)
for i in range(2, n+2):
    sheet['A'+str(i)] = i-1
for i in range(2, n+2):
    sheet.cell(row = 1, column = i).value = i-1

for i in range(1, n+1):
    for y in range(1, n+1):
        sheet.cell(row=i+1, column=y+1).value = i * y


wb.save('multiplicationTable.xlsx')
print('Done.')
    
