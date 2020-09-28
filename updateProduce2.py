import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
print('Updating cells...')
PRICE_UPDATES = {'Celery': 1.19, 'Garlic': 3.07, 'Lemon': 1.27} 
for rowNum in range(2, sheet.max_row+1):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
print('Done.')
